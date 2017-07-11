# -*- encoding:utf-8 -*-


from openpyxl import *

CASE_LIST = []

wb = load_workbook("adtest.xlsx")
print wb.get_sheet_names()

ws = wb.get_sheet_by_name("adtest")

# print "Work Sheet Titile:", ws.title
# print "Work Sheet Rows:", ws.max_row
# print "Work Sheet Cols:", ws.max_column
a = ws['A2'].value
# print a
# print ws.cell(row = 1,column=2).value

#excel 载入 二维list中
for i in range(1, ws.max_row + 1):
    list_temp = []
    for j in range(2, ws.max_column + 1):
        list_temp.append(ws.cell(row=i, column=j).value)
    CASE_LIST.append(list_temp)

print CASE_LIST